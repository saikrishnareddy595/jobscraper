"""
excel_export.py — Generates a rich, multi-sheet Excel workbook after each pipeline run.

Sheets produced:
  1. Jobs              — All scored jobs with visa-sponsorship flag highlighted
  2. Visa Sponsorship  — Only jobs where visa_sponsorship is True (H1B/F1/OPT)
  3. Recruiters        — All discovered recruiters with AI-guessed email callout
  4. Jobs + Recruiters — Joined view: one row per job-recruiter pair

Styling:
  • Frozen header row, bold headers, alternating row colours
  • Visa Sponsorship column colour-coded: green=Yes, red=No, yellow=Unknown
  • AI-guessed email cells highlighted in amber with "(AI Guess)" appended
  • Confidence score column uses a colour gradient bar (conditional format)
  • Auto-fitted column widths
  • All hyperlinks are clickable

Dependencies:
  openpyxl  (already lightweight; add to requirements.txt if not present)

Usage (standalone):
  from output.excel_export import ExcelExporter
  exporter = ExcelExporter()
  path = exporter.export(jobs, recruiter_map)
  print(f"Saved to {path}")
"""

import logging
import os
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# ── Output path ───────────────────────────────────────────────────────────────
_DEFAULT_DIR = os.path.join(os.path.dirname(__file__), "..", "exports")


# ── Colours (openpyxl ARGB hex strings, no leading #) ────────────────────────
_C_HEADER_BG   = "FF1F3864"   # dark navy
_C_HEADER_FG   = "FFFFFFFF"   # white
_C_ROW_ALT     = "FFF2F6FC"   # light blue-grey
_C_ROW_EVEN    = "FFFFFFFF"   # white
_C_VISA_YES    = "FFD6EDD6"   # soft green
_C_VISA_NO     = "FFFFD6D6"   # soft red
_C_VISA_UNK    = "FFFFFFD6"   # soft yellow
_C_AI_EMAIL    = "FFFFF3CD"   # amber
_C_URL         = "FF1155CC"   # hyperlink blue
_C_SCORE_HIGH  = "FFD4EDDA"   # green tint for high score
_C_SCORE_MED   = "FFFFF3CD"   # yellow tint for medium score
_C_SCORE_LOW   = "FFFFD6D6"   # red tint for low score
_C_SECTION_HDR = "FF263859"   # section subheader colour


def _try_import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
        from openpyxl.utils import get_column_letter
        import sqlite3
        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, sqlite3
    except ImportError:
        logger.error(
            "openpyxl not installed — run: pip install openpyxl\n"
            "Excel export disabled."
        )
        return None, None, None, None, None, None, None, None


def _fmt_salary(salary) -> str:
    if salary is None:
        return "N/A"
    try:
        return f"${int(salary):,}"
    except (ValueError, TypeError):
        return str(salary)


def _fmt_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    try:
        dt = datetime.fromisoformat(str(val))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(val)


def _fmt_visa(val: Optional[bool]) -> str:
    if val is True:
        return "✅ Yes"
    if val is False:
        return "❌ No"
    return "❓ Unknown"


def _visa_colour(val: Optional[bool]) -> str:
    if val is True:
        return _C_VISA_YES
    if val is False:
        return _C_VISA_NO
    return _C_VISA_UNK


def _score_colour(score: int) -> str:
    if score >= 70:
        return _C_SCORE_HIGH
    if score >= 40:
        return _C_SCORE_MED
    return _C_SCORE_LOW


class ExcelExporter:
    """
    Generates a fully formatted Excel workbook from the pipeline output.

    Args:
        export_dir: Directory to write the .xlsx file into.
                    Defaults to  job_scraper/exports/
    """

    def __init__(self, export_dir: Optional[str] = None):
        self._export_dir = export_dir or _DEFAULT_DIR
        os.makedirs(self._export_dir, exist_ok=True)

    # ── Public API ─────────────────────────────────────────────────────────────

    def export(
        self,
        jobs: List[Dict[str, Any]],
        recruiter_map: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    ) -> Optional[str]:
        """
        Write the Excel workbook and return the file path.

        Args:
            jobs:           Scored + filtered jobs from the pipeline.
            recruiter_map:  {job_url: [recruiter_dict, ...]} from RecruiterEngine.

        Returns:
            Absolute path to the .xlsx file, or None if openpyxl is unavailable.
        """
        mods = _try_import_openpyxl()
        openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, sqlite3 = mods
        if openpyxl is None:
            return None

        recruiter_map = recruiter_map or {}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove the default blank sheet

        # ----- Sheet 1: All Jobs -----
        self._build_jobs_sheet(wb, jobs, Font, PatternFill, Alignment, get_column_letter)

        # ----- Sheet 2: Visa Sponsorship Jobs -----
        visa_jobs = [j for j in jobs if j.get("visa_sponsorship") is True]
        self._build_visa_sheet(wb, visa_jobs, Font, PatternFill, Alignment, get_column_letter)

        # ----- Sheet 3: Recruiters -----
        all_recruiters = self._flatten_recruiter_map(recruiter_map)
        self._build_recruiters_sheet(
            wb, all_recruiters, Font, PatternFill, Alignment, get_column_letter
        )

        # ----- Sheet 4: Jobs + Recruiters (joined) -----
        self._build_joined_sheet(
            wb, jobs, recruiter_map, Font, PatternFill, Alignment, get_column_letter
        )
        
        # ----- Sheet 5: LinkedIn Posts -----
        try:
            self._build_linkedin_posts_sheet(
                wb, sqlite3, Font, PatternFill, Alignment, get_column_letter
            )
        except Exception as e:
            logger.warning(f"Failed to build LinkedIn posts Excel sheet: {e}")
            
        # ----- Sheet 6: Outreach Messages -----
        self._build_outreach_sheet(
            wb, sqlite3, Font, PatternFill, Alignment, get_column_letter
        )

        # ----- Save -----
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(self._export_dir, f"job_search_{ts}.xlsx")
        wb.save(path)
        logger.info("Excel export saved: %s", path)
        return path

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _build_jobs_sheet(self, wb, jobs, Font, PatternFill, Alignment, get_column_letter):
        ws = wb.create_sheet("All Jobs")
        headers = [
            "Score", "Title", "Company", "Location", "Salary",
            "Visa Sponsorship", "Job Type", "Source",
            "Easy Apply", "Applicants", "Posted Date", "Apply URL",
        ]
        self._write_header(ws, headers, Font, PatternFill, Alignment)

        for i, job in enumerate(jobs, start=2):
            row_bg = _C_ROW_EVEN if i % 2 == 0 else _C_ROW_ALT
            visa   = job.get("visa_sponsorship")

            cells = [
                job.get("score", 0),
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                _fmt_salary(job.get("salary")),
                _fmt_visa(visa),
                (job.get("job_type") or "").replace("_", " ").title(),
                job.get("source", ""),
                "Yes" if job.get("easy_apply") else ("No" if job.get("easy_apply") is False else ""),
                job.get("applicants") or "",
                _fmt_date(job.get("posted_date")),
                job.get("url", ""),
            ]
            ws.append(cells)

            for col_idx, cell in enumerate(ws[i], start=1):
                # Score colouring
                if col_idx == 1:
                    self._fill_cell(cell, PatternFill, _score_colour(job.get("score", 0)))
                # Visa colouring
                elif col_idx == 6:
                    self._fill_cell(cell, PatternFill, _visa_colour(visa))
                else:
                    self._fill_cell(cell, PatternFill, row_bg)
                cell.alignment = Alignment(wrap_text=False, vertical="center")

            # Make URL clickable
            url_cell = ws.cell(row=i, column=12)
            if job.get("url"):
                url_cell.hyperlink = job["url"]
                url_cell.font = Font(color=_C_URL, underline="single")
                url_cell.value = "Apply →"

        self._autofit(ws, get_column_letter, max_width=50)
        ws.freeze_panes = "A2"

    def _build_visa_sheet(self, wb, visa_jobs, Font, PatternFill, Alignment, get_column_letter):
        ws = wb.create_sheet(f"Visa Sponsorship ({len(visa_jobs)})")
        headers = [
            "Score", "Title", "Company", "Location", "Salary",
            "Job Type", "Source", "Easy Apply", "Posted Date", "Apply URL",
        ]
        self._write_header(ws, headers, Font, PatternFill, Alignment)

        for i, job in enumerate(visa_jobs, start=2):
            row_bg = _C_VISA_YES if i % 2 == 0 else "FFD0EAD0"
            cells = [
                job.get("score", 0),
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                _fmt_salary(job.get("salary")),
                (job.get("job_type") or "").replace("_", " ").title(),
                job.get("source", ""),
                "Yes" if job.get("easy_apply") else ("No" if job.get("easy_apply") is False else ""),
                _fmt_date(job.get("posted_date")),
                job.get("url", ""),
            ]
            ws.append(cells)
            for col_idx, cell in enumerate(ws[i], start=1):
                self._fill_cell(cell, PatternFill, _score_colour(job.get("score", 0)) if col_idx == 1 else row_bg)
                cell.alignment = Alignment(wrap_text=False, vertical="center")
            url_cell = ws.cell(row=i, column=10)
            if job.get("url"):
                url_cell.hyperlink = job["url"]
                url_cell.font = Font(color=_C_URL, underline="single")
                url_cell.value = "Apply →"

        # Add a summary note at the top (insert before data rows)
        ws.insert_rows(1)
        note_cell = ws["A1"]
        note_cell.value = (
            f"🎯 {len(visa_jobs)} jobs that explicitly mention H1B / F1 / OPT visa sponsorship.  "
            f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}."
        )
        note_cell.font = Font(italic=True, color="FF1F3864", size=10)
        ws.merge_cells("A1:J1")
        ws.freeze_panes = "A3"
        self._autofit(ws, get_column_letter, max_width=50)

    def _build_recruiters_sheet(
        self, wb, all_recruiters, Font, PatternFill, Alignment, get_column_letter
    ):
        ws = wb.create_sheet("Recruiters")
        headers = [
            "Name", "Title", "Company", "Email", "Email Source",
            "All Email Guesses", "LinkedIn URL", "Location",
            "Confidence Score", "Discovered For (Job URL)",
        ]
        self._write_header(ws, headers, Font, PatternFill, Alignment)

        for i, rec in enumerate(all_recruiters, start=2):
            is_ai   = rec.get("email_is_ai_guess", False)
            email   = rec.get("email", "")
            guesses = "; ".join(rec.get("guessed_emails", []))
            score   = rec.get("confidence_score", 0)

            cells = [
                rec.get("name", ""),
                rec.get("title", ""),
                rec.get("company", ""),
                email + (" (AI Guess ⚠)" if is_ai else ""),
                "🤖 AI Generated" if is_ai else "✅ Found in profile",
                guesses,
                rec.get("linkedin_url", ""),
                rec.get("location", ""),
                score,
                rec.get("job_url", ""),
            ]
            ws.append(cells)

            row_bg = _C_ROW_EVEN if i % 2 == 0 else _C_ROW_ALT
            for col_idx, cell in enumerate(ws[i], start=1):
                # Email column: amber if AI guessed
                if col_idx == 4 and is_ai:
                    self._fill_cell(cell, PatternFill, _C_AI_EMAIL)
                    cell.font = Font(color="FF856404")
                elif col_idx == 9:
                    self._fill_cell(cell, PatternFill, _score_colour(score))
                else:
                    self._fill_cell(cell, PatternFill, row_bg)
                cell.alignment = Alignment(wrap_text=False, vertical="center")

            # Clickable LinkedIn URL
            li_cell = ws.cell(row=i, column=7)
            if rec.get("linkedin_url"):
                li_cell.hyperlink = rec["linkedin_url"]
                li_cell.font = Font(color=_C_URL, underline="single")
                li_cell.value = rec["name"] or "LinkedIn Profile"

        self._autofit(ws, get_column_letter, max_width=55)
        ws.freeze_panes = "A2"

    def _build_joined_sheet(
        self, wb, jobs, recruiter_map, Font, PatternFill, Alignment, get_column_letter
    ):
        ws = wb.create_sheet("Jobs + Recruiters")
        headers = [
            "Job Score", "Job Title", "Company", "Location", "Salary",
            "Visa Sponsorship", "Recruiter Name", "Recruiter Title",
            "Recruiter Email", "Email Source", "LinkedIn URL",
            "Confidence", "Apply URL",
        ]
        self._write_header(ws, headers, Font, PatternFill, Alignment)

        row_num = 2
        for job in jobs:
            url = job.get("url", "")
            recruiters = recruiter_map.get(url, [])
            visa = job.get("visa_sponsorship")

            if not recruiters:
                # Job with no recruiter found — still show the job row
                row = self._joined_row(job, {}, visa)
                ws.append(row)
                self._style_joined_row(ws, row_num, job, {}, PatternFill, Font, Alignment, visa)
                row_num += 1
            else:
                for rec in recruiters:
                    row = self._joined_row(job, rec, visa)
                    ws.append(row)
                    self._style_joined_row(ws, row_num, job, rec, PatternFill, Font, Alignment, visa)
                    row_num += 1

        self._autofit(ws, get_column_letter, max_width=50)
        ws.freeze_panes = "A2"

    def _joined_row(self, job, rec, visa) -> list:
        is_ai = rec.get("email_is_ai_guess", False)
        email = rec.get("email", "")
        if is_ai and email:
            email_display = email + " (AI Guess ⚠)"
        else:
            email_display = email

        return [
            job.get("score", 0),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            _fmt_salary(job.get("salary")),
            _fmt_visa(visa),
            rec.get("name", ""),
            rec.get("title", ""),
            email_display,
            "🤖 AI" if is_ai else ("✅ Real" if email else ""),
            rec.get("linkedin_url", ""),
            rec.get("confidence_score", ""),
            job.get("url", ""),
        ]

    def _style_joined_row(self, ws, row_num, job, rec, PatternFill, Font, Alignment, visa):
        row_bg = _C_ROW_EVEN if row_num % 2 == 0 else _C_ROW_ALT
        is_ai  = rec.get("email_is_ai_guess", False)
        for col_idx, cell in enumerate(ws[row_num], start=1):
            if col_idx == 1:
                self._fill_cell(cell, PatternFill, _score_colour(job.get("score", 0)))
            elif col_idx == 6:
                self._fill_cell(cell, PatternFill, _visa_colour(visa))
            elif col_idx == 9 and is_ai:
                self._fill_cell(cell, PatternFill, _C_AI_EMAIL)
                cell.font = Font(color="FF856404")
            elif col_idx == 12 and rec.get("confidence_score"):
                self._fill_cell(cell, PatternFill, _score_colour(rec.get("confidence_score", 0)))
            else:
                self._fill_cell(cell, PatternFill, row_bg)
            cell.alignment = Alignment(wrap_text=False, vertical="center")

        # Hyperlinks
        li_cell = ws.cell(row=row_num, column=11)
        if rec.get("linkedin_url"):
            li_cell.hyperlink = rec["linkedin_url"]
            li_cell.font = Font(color=_C_URL, underline="single")
            li_cell.value = rec.get("name", "LinkedIn Profile") or "LinkedIn Profile"

        url_cell = ws.cell(row=row_num, column=13)
        if job.get("url"):
            url_cell.hyperlink = job["url"]
            url_cell.font = Font(color=_C_URL, underline="single")
            url_cell.value = "Apply →"

    def _build_outreach_sheet(self, wb, sqlite3, Font, PatternFill, Alignment, get_column_letter):
        ws = wb.create_sheet("Outreach Drafts")
        headers = [
            "Job ID", "Recruiter ID", "Message Type", "Tone", "Subject", "Body", "Fit Score", "Status", "Generated At"
        ]
        self._write_header(ws, headers, Font, PatternFill, Alignment)
        
        try:
            import config
            conn = sqlite3.connect(config.DB_PATH)
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT * FROM outreach_messages ORDER BY created_at DESC LIMIT 500").fetchall()
            conn.close()
        except Exception as e:
            logger.warning(f"Could not load outreach messages for excel export: {e}")
            rows = []
            
        for i, row in enumerate(rows, start=2):
            row_bg = _C_ROW_EVEN if i % 2 == 0 else _C_ROW_ALT
            cells = [
                row["job_id"],
                row["recruiter_id"],
                row["message_type"],
                row["tone"],
                row["subject"],
                row["body"],
                row["fit_score"],
                row["status"],
                row["created_at"]
            ]
            ws.append(cells)
            for col_idx, cell in enumerate(ws[i], start=1):
                self._fill_cell(cell, PatternFill, row_bg)
                cell.alignment = Alignment(wrap_text=True if col_idx == 6 else False, vertical="center")
                
        self._autofit(ws, get_column_letter, max_width=60)
        # Manually stretch the body column layout
        ws.column_dimensions[get_column_letter(6)].width = 80
        ws.freeze_panes = "A2"

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _write_header(self, ws, headers, Font, PatternFill, Alignment):
        ws.append(headers)
        header_row = ws[1]
        for cell in header_row:
            cell.font      = Font(bold=True, color=_C_HEADER_FG, size=11)
            cell.fill      = PatternFill(fill_type="solid", fgColor=_C_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

    @staticmethod
    def _fill_cell(cell, PatternFill, colour: str):
        cell.fill = PatternFill(fill_type="solid", fgColor=colour)

    @staticmethod
    def _autofit(ws, get_column_letter, max_width: int = 40):
        """Approximate column width based on max content length."""
        col_widths: Dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    length = len(str(cell.value))
                    col_widths[cell.column] = max(col_widths.get(cell.column, 0), length)
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, max_width)

    @staticmethod
    def _flatten_recruiter_map(
        recruiter_map: Dict[str, List[Dict[str, Any]]],
    ) -> List[Dict[str, Any]]:
        """Flatten recruiter_map into a flat list, attaching job_url to each record."""
        seen: set = set()
        result: List[Dict[str, Any]] = []
        for job_url, recs in recruiter_map.items():
            for rec in recs:
                key = rec.get("linkedin_url") or (rec.get("name", "") + rec.get("company", ""))
                if key and key in seen:
                    continue
                seen.add(key)
                result.append({**rec, "job_url": job_url})
        result.sort(key=lambda r: r.get("confidence_score", 0), reverse=True)
        return result

    def _build_linkedin_posts_sheet(self, wb, sqlite3, Font, PatternFill, Alignment, get_column_letter):
        ws = wb.create_sheet("LinkedIn Post Leads")
        headers = [
            "Score", "Extracted Title", "Extracted Company", "Author Name", "Headline",
            "Contact Email", "Contact LinkedIn", "Post Text", "Post URL", "Role Category", "Date Scraped"
        ]
        self._write_header(ws, headers, Font, PatternFill, Alignment)
        
        import config
        try:
            conn = sqlite3.connect(config.DB_PATH)
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT * FROM linkedin_posts ORDER BY score DESC, scraped_at DESC LIMIT 500").fetchall()
            conn.close()
        except:
            rows = []
            
        for i, row in enumerate(rows, start=2):
            row_bg = _C_ROW_EVEN if i % 2 == 0 else _C_ROW_ALT
            cells = [
                row["score"],
                row["extracted_title"],
                row["extracted_company"],
                row["author_name"],
                row["author_headline"],
                row["contact_email"],
                row["contact_linkedin"],
                row["post_text"][:2000] if row["post_text"] else "",
                row["post_url"],
                row["role_category"],
                row["scraped_at"]
            ]
            ws.append(cells)
            for col_idx, cell in enumerate(ws[i], start=1):
                if col_idx == 1:
                    self._fill_cell(cell, PatternFill, _score_colour(row["score"]))
                else:
                    self._fill_cell(cell, PatternFill, row_bg)
                cell.alignment = Alignment(wrap_text=True if col_idx == 8 else False, vertical="center")
                
            # Links
            url_cell = ws.cell(row=i, column=9)
            if row["post_url"]:
                url_cell.hyperlink = row["post_url"]
                url_cell.font = Font(color=_C_URL, underline="single")
                url_cell.value = "View Post →"
                
        self._autofit(ws, get_column_letter, max_width=45)
        ws.column_dimensions[get_column_letter(8)].width = 75
        ws.freeze_panes = "A2"
